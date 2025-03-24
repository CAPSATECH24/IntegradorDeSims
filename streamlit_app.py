import openpyxl 
import pandas as pd
import sqlite3
import re
import os
import streamlit as st
import logging
from io import BytesIO
from datetime import date

# Configuración básica de logging
logging.basicConfig(level=logging.INFO, filename='procesamiento.log', filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Nombre del archivo de salida con la fecha de hoy
output_filename = f"dei Sims ({date.today().strftime('%Y-%m-%d')}).db"

# Mapeos predeterminados basados en el nombre de la pestaña
default_mappings = {
    "SIMPATIC": {
        'ICCID': 'ICCID',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'Fecha Vencimiento',
        'EN SESION': 'Fecha Vencimiento',
        'ConsumoMb': 'Fecha Vencimiento'
    },
    "TELCEL ALEJANDRO": {
        'ICCID': 'ICCID',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'ESTADO SIM',
        'EN SESION': 'SESIÓN',
        'ConsumoMb': 'LÍMITE DE USO DE DATOS' 
    },
    "-1": {
        'ICCID': 'ICCID',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'Estado de SIM',
        'EN SESION': 'En sesión',
        'ConsumoMb': 'Uso de ciclo hasta la fecha (MB)'  
    },
    "-2": {
        'ICCID': 'ICCID',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'Estado de SIM',
        'EN SESION': 'En sesión',
        'ConsumoMb': 'Uso de ciclo hasta la fecha (MB)'  
    },
    "TELCEL": {
        'ICCID': 'Cuenta Padre',
        'TELEFONO': 'Línea',
        'ESTADO DEL SIM': 'Estatus línea',
        'EN SESION': 'Estatus línea',
        'ConsumoMb': 'Motivo línea' 
    },
    "MOVISTAR": {
        'ICCID': 'ICC',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'Estado',
        'EN SESION': 'Estado GPRS',
        'ConsumoMb': 'Consumo Datos Mensual' 
    },
    "NANTI": {
        'ICCID': 'ICCID',
        'TELEFONO': 'MSISDN',
        'ESTADO DEL SIM': 'Estado',
        'EN SESION': 'Estado',
        'ConsumoMb': 'Estado'
    },
    "LEGACY": {
        'ICCID': 'ICCID',
        'TELEFONO': 'normalized_key',
        'ESTADO DEL SIM': 'ESTADO_DEL_SIM',
        'EN SESION': 'EN_SESION',
        'ConsumoMb': 'ConsumoMb'
    }
}

def create_database(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(''' 
        CREATE TABLE IF NOT EXISTS sims ( 
            ICCID TEXT, 
            TELEFONO TEXT, 
            ESTADO_DEL_SIM TEXT, 
            EN_SESION TEXT, 
            ConsumoMb TEXT,
            Compania TEXT,
            UNIQUE(ICCID, TELEFONO)
        ) 
    ''')
    conn.commit()
    conn.close()

# Función para insertar datos en la base de datos con manejo de duplicados
def insert_data(db_path, data):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    records_before = cursor.execute("SELECT COUNT(*) FROM sims").fetchone()[0]
    try:
        cursor.executemany(
            "INSERT OR IGNORE INTO sims (ICCID, TELEFONO, ESTADO_DEL_SIM, EN_SESION, ConsumoMb, Compania) VALUES (?, ?, ?, ?, ?, ?)",
            data
        )
        conn.commit()
        records_after = cursor.execute("SELECT COUNT(*) FROM sims").fetchone()[0]
        records_inserted = records_after - records_before
        logging.info(f"Insertados {records_inserted} registros en la base de datos.")
        return len(data), records_inserted
    finally:
        conn.close()

# Función para limpiar ICCID, TELEFONO y ConsumoMb manteniendo ceros a la izquierda
def clean_iccid_telefono_consumo(data):
    cleaned_data = []
    for row in data:
        cleaned_row = list(row)
        original_iccid = cleaned_row[0]
        original_telefono = cleaned_row[1]
        original_consumo_mb = cleaned_row[4]
        
        # Limpieza de ICCID
        iccid_value = row[0]
        if isinstance(iccid_value, float) and iccid_value.is_integer():
            cleaned_iccid = str(int(iccid_value))
        else:
            cleaned_iccid = str(iccid_value)
        cleaned_row[0] = ''.join(filter(str.isdigit, cleaned_iccid)) if cleaned_iccid else ""
        
        # Limpieza de TELEFONO
        telefono_value = row[1]
        if isinstance(telefono_value, float) and telefono_value.is_integer():
            cleaned_telefono = str(int(telefono_value))
        else:
            cleaned_telefono = str(telefono_value)
        cleaned_row[1] = ''.join(filter(str.isdigit, cleaned_telefono)) if cleaned_telefono else ""
        
        # Limpieza de ConsumoMb
        consumo_mb_value = row[4]
        cleaned_consumo_mb = ''.join(filter(str.isdigit, consumo_mb_value)) if consumo_mb_value else ""
        cleaned_row[4] = cleaned_consumo_mb
        
        # Normalizar otros campos
        cleaned_row[2] = cleaned_row[2].strip().lower() if cleaned_row[2] else ""
        cleaned_row[3] = cleaned_row[3].strip().lower() if cleaned_row[3] else ""
        
        cleaned_data.append(tuple(cleaned_row))
        logging.info(f"Limpieza Registro: ICCID '{original_iccid}' a '{cleaned_row[0]}', TELEFONO '{original_telefono}' a '{cleaned_row[1]}', ConsumoMb '{original_consumo_mb}' a '{cleaned_row[4]}'")
    return cleaned_data

# Función para normalizar otros campos
def normalize_data(data):
    normalized_data = []
    for row in data:
        cleaned_row = list(row)
        cleaned_row[2] = cleaned_row[2].strip().lower() if cleaned_row[2] else ""
        cleaned_row[3] = cleaned_row[3].strip().lower() if cleaned_row[3] else ""
        normalized_data.append(tuple(cleaned_row))
    return normalized_data

# Función para procesar archivos Excel (usando file bytes)
def process_excel(file_bytes, column_mapping, sheet_name):
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook[sheet_name]
    all_data = []
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            row_data = []
            for key in ['ICCID', 'TELEFONO', 'ESTADO DEL SIM', 'EN SESION', 'ConsumoMb']:
                col_index = column_mapping[key]
                if col_index is None or col_index == -1:
                    cell_value = ""
                elif col_index >= len(row):
                    cell_value = ""
                else:
                    cell = row[col_index]
                    if isinstance(cell, float) and cell.is_integer():
                        cell_value = str(int(cell))
                    elif isinstance(cell, (int, str)):
                        cell_value = str(cell)
                    else:
                        cell_value = str(cell) if cell is not None else ""
                row_data.append(cell_value)
            row_data.append(sheet_name)  # Añadir el nombre de la pestaña como 'Compania'
            all_data.append(row_data)
        except IndexError:
            st.warning(f"Error procesando fila en la pestaña '{sheet_name}'. Fila omitida.")
    return all_data

# Función para procesar archivos CSV (usando file bytes)
def process_csv(file_bytes, column_mapping):
    try:
        df = pd.read_csv(BytesIO(file_bytes), dtype=str)
    except Exception as e:
        logging.error(f"Error leyendo CSV: {e}")
        return []
    all_data = []
    company_name = "CSV"  # Se puede ajustar si se desea usar otro identificador
    for index, row in df.iterrows():
        try:
            row_data = []
            for key in ['ICCID', 'TELEFONO', 'ESTADO DEL SIM', 'EN SESION', 'ConsumoMb']:
                cell = row.get(column_mapping[key], "")
                if pd.notnull(cell):
                    cell = cell.strip()
                    if re.match(r'^\d+\.\0+$', cell):
                        cell_value = str(int(float(cell)))
                    else:
                        cell_value = re.sub(r'[^\d]', '', cell)
                else:
                    cell_value = ""
                row_data.append(cell_value)
            row_data.append(company_name)
            all_data.append(row_data)
        except KeyError:
            st.warning(f"Error procesando fila {index + 1} en el archivo CSV. Fila omitida.")
    return all_data

# Función auxiliar para permitir la selección manual de columnas
def get_column_selection(columns, label, key):
    selection = st.selectbox(
        label,
        options=columns,
        index=0,
        key=key
    )
    return selection

# Interfaz de usuario con Streamlit
st.title("Carga de Excel y CSV y Homologación de Base de Datos")

# Permitir que el usuario suba archivos directamente
uploaded_files = st.file_uploader("Carga los archivos Excel o CSV:", accept_multiple_files=True, type=["xlsx", "csv"])

if uploaded_files:
    st.write(f"Archivos cargados: {[file.name for file in uploaded_files]}")
    column_mapping = {}  # Almacenará el mapeo para cada archivo

    # Para cada archivo cargado se crea un mapeo de columnas
    for file in uploaded_files:
        file_bytes = file.getvalue()
        sheet_data = {}
        if file.name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
            for sheet_name in workbook.sheetnames:
                st.header(f"Archivo: {file.name} | Pestaña: {sheet_name}")
                header_row = next(workbook[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True))
                if sheet_name in default_mappings:
                    mapping = default_mappings[sheet_name]
                    mapping_indices = {}
                    mapping_valid = True
                    for key_field, column_name in mapping.items():
                        if column_name in header_row:
                            mapping_indices[key_field] = header_row.index(column_name)
                        else:
                            st.warning(f"La columna '{column_name}' no se encontró en la pestaña '{sheet_name}' del archivo '{file.name}'. Se requiere selección manual.")
                            mapping_valid = False
                            break
                    if mapping_valid:
                        if 'ConsumoMb' in mapping:
                            if mapping['ConsumoMb'] in header_row:
                                mapping_indices['ConsumoMb'] = header_row.index(mapping['ConsumoMb'])
                            else:
                                mapping_indices['ConsumoMb'] = -1
                        sheet_data[sheet_name] = mapping_indices
                        st.info(f"Usando mapeo predeterminado para la pestaña '{sheet_name}' del archivo '{file.name}'.")
                        logging.info(f"Archivo: {file.name} | Pestaña: {sheet_name} | Mapeo: {mapping_indices}")
                    else:
                        columns = [cell if cell is not None else "" for cell in header_row]
                        st.write("Selecciona las columnas correspondientes para cada campo requerido:")
                        iccid_col = get_column_selection(columns, label="Selecciona columna para ICCID:", key=f"{file.name}_{sheet_name}_iccid")
                        telefono_col = get_column_selection(columns, label="Selecciona columna para TELEFONO:", key=f"{file.name}_{sheet_name}_telefono")
                        estado_sim_col = get_column_selection(columns, label="Selecciona columna para ESTADO DEL SIM:", key=f"{file.name}_{sheet_name}_estado_sim")
                        en_sesion_col = get_column_selection(columns, label="Selecciona columna para EN SESION:", key=f"{file.name}_{sheet_name}_en_sesion")
                        consumo_mb_col = get_column_selection(columns, label="Selecciona columna para ConsumoMb:", key=f"{file.name}_{sheet_name}_consumo_mb")
                        sheet_data[sheet_name] = {
                            'ICCID': columns.index(iccid_col),
                            'TELEFONO': columns.index(telefono_col),
                            'ESTADO DEL SIM': columns.index(estado_sim_col),
                            'EN SESION': columns.index(en_sesion_col),
                            'ConsumoMb': columns.index(consumo_mb_col) if consumo_mb_col in columns else -1
                        }
                        logging.info(f"Archivo: {file.name} | Pestaña: {sheet_name} | Mapeo Manual: {sheet_data[sheet_name]}")
                else:
                    columns = [cell if cell is not None else "" for cell in next(workbook[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True))]
                    st.write("Selecciona las columnas correspondientes para cada campo requerido:")
                    iccid_col = get_column_selection(columns, label="Selecciona columna para ICCID:", key=f"{file.name}_{sheet_name}_iccid")
                    telefono_col = get_column_selection(columns, label="Selecciona columna para TELEFONO:", key=f"{file.name}_{sheet_name}_telefono")
                    estado_sim_col = get_column_selection(columns, label="Selecciona columna para ESTADO DEL SIM:", key=f"{file.name}_{sheet_name}_estado_sim")
                    en_sesion_col = get_column_selection(columns, label="Selecciona columna para EN SESION:", key=f"{file.name}_{sheet_name}_en_sesion")
                    consumo_mb_col = get_column_selection(columns, label="Selecciona columna para ConsumoMb:", key=f"{file.name}_{sheet_name}_consumo_mb")
                    sheet_data[sheet_name] = {
                        'ICCID': columns.index(iccid_col),
                        'TELEFONO': columns.index(telefono_col),
                        'ESTADO DEL SIM': columns.index(estado_sim_col),
                        'EN SESION': columns.index(en_sesion_col),
                        'ConsumoMb': columns.index(consumo_mb_col) if consumo_mb_col in columns else -1
                    }
                    logging.info(f"Archivo: {file.name} | Pestaña: {sheet_name} | Mapeo Manual: {sheet_data[sheet_name]}")
        elif file.name.endswith('.csv'):
            df = pd.read_csv(BytesIO(file.getvalue()), dtype=str)
            columns = df.columns.tolist()
            st.header(f"Archivo: {file.name}")
            st.write("Selecciona las columnas correspondientes para cada campo requerido:")
            iccid_col = get_column_selection(columns, label="Selecciona columna para ICCID:", key=f"{file.name}_iccid")
            telefono_col = get_column_selection(columns, label="Selecciona columna para TELEFONO:", key=f"{file.name}_telefono")
            estado_sim_col = get_column_selection(columns, label="Selecciona columna para ESTADO DEL SIM:", key=f"{file.name}_estado_sim")
            en_sesion_col = get_column_selection(columns, label="Selecciona columna para EN SESION:", key=f"{file.name}_en_sesion")
            consumo_mb_col = get_column_selection(columns, label="Selecciona columna para ConsumoMb:", key=f"{file.name}_consumo_mb")
            sheet_data[file.name] = {
                'ICCID': df.columns.get_loc(iccid_col),
                'TELEFONO': df.columns.get_loc(telefono_col),
                'ESTADO DEL SIM': df.columns.get_loc(estado_sim_col),
                'EN SESION': df.columns.get_loc(en_sesion_col),
                'ConsumoMb': df.columns.get_loc(consumo_mb_col) if consumo_mb_col in df.columns else -1
            }
            logging.info(f"Archivo: {file.name} | CSV | Mapeo Manual: {sheet_data[file.name]}")
        column_mapping[file.name] = sheet_data

    # Vista previa del mapeo de columnas
    st.subheader("Vista Previa de Mapeo de Columnas")
    for file in uploaded_files:
        file_bytes = file.getvalue()
        if file.name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
            for sheet, mapping in column_mapping[file.name].items():
                st.write(f"**Archivo:** {file.name} | **Pestaña:** {sheet}")
                header_row = next(workbook[sheet].iter_rows(min_row=1, max_row=1, values_only=True))
                st.write(f" - ICCID: {header_row[mapping['ICCID']]}")
                st.write(f" - TELEFONO: {header_row[mapping['TELEFONO']]}")
                st.write(f" - ESTADO DEL SIM: {header_row[mapping['ESTADO DEL SIM']]}")
                st.write(f" - EN SESION: {header_row[mapping['EN SESION']]}")
                if mapping.get('ConsumoMb', -1) != -1:
                    st.write(f" - ConsumoMb: {header_row[mapping['ConsumoMb']]}")
                else:
                    st.write(" - ConsumoMb: No mapeado")
        elif file.name.endswith('.csv'):
            df = pd.read_csv(BytesIO(file.getvalue()), dtype=str)
            mapping = column_mapping[file.name]
            st.write(f"**Archivo CSV:** {file.name}")
            st.write(f" - ICCID: {df.columns[mapping['ICCID']]}")
            st.write(f" - TELEFONO: {df.columns[mapping['TELEFONO']]}")
            st.write(f" - ESTADO DEL SIM: {df.columns[mapping['ESTADO DEL SIM']]}")
            st.write(f" - EN SESION: {df.columns[mapping['EN SESION']]}")
            if mapping.get('ConsumoMb', -1) != -1:
                st.write(f" - ConsumoMb: {df.columns[mapping['ConsumoMb']]}")
            else:
                st.write(" - ConsumoMb: No mapeado")
    
    # Botón para procesar todos los archivos subidos
    if st.button("Procesar Todos los Archivos"):
        all_mappings_valid = True
        stats_by_file = {}
        total_records = 0
        total_inserted = 0

        # Validación de mapeos
        for file in uploaded_files:
            file_bytes = file.getvalue()
            if file.name.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
                stats_by_file[file.name] = {'sheets': {}}
                for sheet_name in workbook.sheetnames:
                    num_columns = workbook[sheet_name].max_column
                    for key in ['ICCID', 'TELEFONO', 'ESTADO DEL SIM', 'EN SESION', 'ConsumoMb']:
                        if key == 'ConsumoMb' and column_mapping[file.name][sheet_name][key] == -1:
                            st.warning(f"La columna 'ConsumoMb' no está mapeada para la pestaña '{sheet_name}' del archivo '{file.name}'. Se establecerá como NULL.")
                        elif key != 'ConsumoMb' and not (0 <= column_mapping[file.name][sheet_name][key] < num_columns):
                            st.error(f"Mapeo inválido para '{key}' en la pestaña '{sheet_name}' del archivo '{file.name}'.")
                            all_mappings_valid = False
            elif file.name.endswith('.csv'):
                df = pd.read_csv(BytesIO(file.getvalue()), dtype=str)
                mapping = column_mapping[file.name]
                num_columns = len(df.columns)
                for key in ['ICCID', 'TELEFONO', 'ESTADO DEL SIM', 'EN SESION', 'ConsumoMb']:
                    if key == 'ConsumoMb' and mapping[key] == -1:
                        st.warning(f"La columna 'ConsumoMb' no está mapeada para el archivo CSV '{file.name}'. Se establecerá como NULL.")
                    elif key != 'ConsumoMb' and not (0 <= mapping[key] < num_columns):
                        st.error(f"Mapeo inválido para '{key}' en el archivo CSV '{file.name}'.")
                        all_mappings_valid = False

        if not all_mappings_valid:
            st.error("Por favor, revisa los mapeos de columnas y corrige los errores antes de proceder.")
        else:
            # Elimina el archivo de BD si existe para que cada ejecución sea independiente
            if os.path.exists(output_filename):
                os.remove(output_filename)
                logging.info(f"Archivo existente {output_filename} eliminado para nueva ejecución.")
            
            create_database(output_filename)
            logging.info(f"Base de datos creada: {output_filename}")

            for file in uploaded_files:
                file_bytes = file.getvalue()
                if file.name.endswith('.xlsx'):
                    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
                    stats_by_file[file.name] = {'sheets': {}}
                    for sheet_name in workbook.sheetnames:
                        data = process_excel(file_bytes, column_mapping[file.name][sheet_name], sheet_name)
                        if data:
                            processed, inserted = insert_data(output_filename, data)
                            stats_by_file[file.name]['sheets'][sheet_name] = {
                                'processed': processed,
                                'inserted': inserted
                            }
                            total_records += processed
                            total_inserted += inserted
                elif file.name.endswith('.csv'):
                    data = process_csv(file_bytes, column_mapping[file.name])
                    if data:
                        processed, inserted = insert_data(output_filename, data)
                        stats_by_file[file.name] = {
                            'processed': processed,
                            'inserted': inserted
                        }
                        total_records += processed
                        total_inserted += inserted

            # Mostrar resultados en pestañas
            tab1, tab2 = st.tabs(["Proceso", "Estadísticas"])
            
            with tab1:
                st.success("¡Procesamiento completado!")
                st.write(f"Total de registros procesados: {total_records}")
                st.write(f"Total de registros insertados: {total_inserted}")
                
            with tab2:
                st.header("Estadísticas de Procesamiento")
                if total_records > 0:
                    st.write(f"Tasa de inserción total: {(total_inserted/total_records*100):.2f}%")
                for file_name, stats in stats_by_file.items():
                    st.subheader(f"Archivo: {file_name}")
                    if 'sheets' in stats:  # Archivo Excel
                        for sheet, sheet_stats in stats['sheets'].items():
                            processed = sheet_stats['processed']
                            inserted = sheet_stats['inserted']
                            insertion_rate = (inserted/processed*100) if processed > 0 else 0
                            st.write(f"Pestaña: {sheet}")
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Registros Procesados", processed)
                            with col2:
                                st.metric("Registros Insertados", inserted)
                            with col3:
                                st.metric("Tasa de Inserción", f"{insertion_rate:.2f}%")
                    else:  # Archivo CSV
                        processed = stats['processed']
                        inserted = stats['inserted']
                        insertion_rate = (inserted/processed*100) if processed > 0 else 0
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Registros Procesados", processed)
                        with col2:
                            st.metric("Registros Insertados", inserted)
                        with col3:
                            st.metric("Tasa de Inserción", f"{insertion_rate:.2f}%")
            
            # Se ofrece para descarga la base de datos generada
            with open(output_filename, "rb") as f:
                db_bytes = f.read()
            st.download_button(
                label="Descargar Base de Datos",
                data=db_bytes,
                file_name=output_filename,
                mime="application/octet-stream"
            )
else:
    st.error("Por favor, sube al menos un archivo Excel o CSV.")
